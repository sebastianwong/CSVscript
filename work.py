from openpyxl import Workbook, load_workbook
import csv

wb1 = load_workbook('FY22_Spending.xlsx')
ws1 = wb1.active

wb2 = load_workbook('FRSE TR.xlsx')
ws2 = wb2.active
cell_obj2 = ws2.cell(row = 1, column = 1)

cell_obj1 = ws1.cell(row = 1, column = 1)

file = open('FY22_Spending.csv')
reader = csv.reader(file)
header = []
header = next(reader)
rows = []
for row in reader:
    rows.append(row)

print(rows)
file.close()


print(cell_obj2.value)
print(cell_obj1.value)